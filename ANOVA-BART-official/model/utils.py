from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

def initialize_excel(config: dict):
    filename = config['experiment_name'] + '_log'
    xlsx_filename = os.path.join(config['experiment_name'], f'{filename}.xlsx')
    num_blocks = config['max_ensembles']
    subfields = ['variable', 'b', 'gamma'] * config['max_depth']
    subfields.extend(['height', 'update'])
    num_subfields = len(subfields)
    total_fields = num_blocks * num_subfields

    wb = Workbook()
    ws = wb.active
    ws.title = 'log'

    ws.cell(row=1, column=1, value = 'train rmse')
    ws.cell(row=2, column=1, value = '')
    ws.cell(row=1, column=2, value = 'test rmse')
    ws.cell(row=2, column=2, value = '')
    ws.cell(row=1, column=3, value = 'error_var')
    ws.cell(row=2, column=3, value = '')
    ws.cell(row=1, column=4, value = 'T')
    ws.cell(row=2, column=4, value = '')

    for block_idx in range(num_blocks):
        start_col = block_idx * num_subfields + 5
        end_col = start_col + num_subfields - 1
        cell = ws.cell(row=1, column=start_col, value=f'tree{block_idx}')
        cell.alignment = Alignment(horizontal = 'center', vertical='center')
        if num_subfields > 1 :
            ws.merge_cells(start_row = 1, start_column=start_col, end_row=1, end_column=end_col)
    
    for col_idx in range(total_fields):
        subfield = subfields[col_idx % num_subfields]
        ws.cell(row = 2, column=col_idx+5, value=subfield)
    
    wb.save(xlsx_filename)


def append_row_from_string(data_strs: list[str], config: dict):
    num_blocks = config['max_ensembles']; num_subfields = 3*config['max_depth']+2
    filename = config['experiment_name'] + '_log'
    xlsx_filename = os.path.join(config['experiment_name'], f'{filename}.xlsx')
    wb = load_workbook(xlsx_filename)
    ws = wb.active
    
    for row_idx, data_str in enumerate(data_strs):
        data = data_str.strip().split(',')
        rmse = data[0]; test_rmse = data[1]; error_var = data[2]; T = data[3]
        data = data[4:]
        if len(data) < num_blocks*num_subfields:
            plus = num_blocks*num_subfields - len(data)
            data.extend(['']*plus)

        ws.cell(row=row_idx+3, column=1, value = rmse)
        ws.cell(row=row_idx+3, column=2, value = test_rmse)
        ws.cell(row=row_idx+3, column=3, value = error_var)
        ws.cell(row=row_idx+3, column=4, value = T)
        for i, value in enumerate(data):
            ws.cell(row=row_idx+3, column = i+5, value=value)

    wb.save(xlsx_filename)
